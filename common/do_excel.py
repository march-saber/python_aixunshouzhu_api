import openpyxl

"""读取写入Excel"""
class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.result = None
        self.sql = None

class DoExcel:
    def __init__(self,file_name,sheet_name):
        try:
            self.file_name = file_name
            self.sheet_name = sheet_name
            self.workbook = openpyxl.load_workbook(file_name)
            self.sheet = self.workbook[sheet_name]
        except Exception as e :
            print("case文件有误：{}".format(e))

    def get_case(self):   #读取文件
        max_row = self.sheet.max_row  #获取最大行

        cases = []
        for r in range(2,max_row+1):   #遍历行数
            # case = {}
            # case["case_id"] = self.sheet.cell(row=r,column=1).value
            # case["title"] = self.sheet.cell(row=r, column=2).value
            # cases.append(case)
            case = Case()   #实例化获取case表的值
            case.case_id = self.sheet.cell(row = r, column = 1).value
            case.title = self.sheet.cell(row=r, column=2).value
            case.url = self.sheet.cell(row=r, column=3).value
            case.data = self.sheet.cell(row=r, column=4).value
            case.method = self.sheet.cell(row=r, column=5).value
            case.expected = self.sheet.cell(row=r, column=6).value
            case.sql = self.sheet.cell(row=r, column=9).value
            cases.append(case)
            self.workbook.close()
        return cases

    def write_result(self,row,actual,result):  #固定行数，写入值
        sheet = self.workbook[self.sheet_name]  #确认sheet
        sheet.cell(row,7).value = actual
        sheet.cell(row,8).value = result
        self.workbook.save(filename=self.file_name)
        self.workbook.close()


