import openpyxl
class DoExcel:
    def __init__(self,excel_name,sheet_name):
        self.workbook = openpyxl.load_workbook(excel_name)
        self.sheet = self.workbook[sheet_name]

    def get_case(self):
        max_row = self.sheet.max_row  # 最大行数
        # max_col = self.sheet.max_column  #最大列数

        # case = []
        case_set = set()
        for r in range(2,max_row+1):
            case_time = self.sheet.cell(r,1).value
            # case_data = self.sheet.cell(r,2).value
            case_set.add(case_time)  #去重的时间
        return case_set

    def get_time(self):
        max_row = self.sheet.max_row
        case = []
        for r in range(2,max_row+1):
            case_time = self.sheet.cell(r,1).value
            case.append(case_time)  #所有时间
        return case


if __name__ == '__main__':
    from openpyxl import load_workbook
    from configparser import ConfigParser

    cf = ConfigParser()
    cf.read('wifi.conf', encoding='utf-8')
    excel = cf.get('wifi_data', 'excel')
    url = "C:\\Users\\t\Desktop\WiFi-data\\time_data\\" + excel
    print(url)
    do_excel1 = DoExcel(url, 'old')
    do_excel2 = DoExcel(url, 'new')
    #去重时间
    old_set = list(do_excel1.get_case())
    new_set = list(do_excel2.get_case())
    #所有时间
    old_case = do_excel1.get_time()

    wb = load_workbook(url)
    sheet3 = wb['old_repetition']
    # sheet4 = wb['new_repetition']

    a = 1
    for item in old_set:
        a+=1
        sheet3.cell(1,1).value = "老版本时间去重复"
        sheet3.cell(a,1).value = item
    wb.save(url)

    # for item in new_set:
    #     a+=1
    #     sheet4.cell(a,1).value = item

    # b = 0
    # for i in old_set:
    #     b +=1
    #     if i in old_case:
    #         old_data = do_excel1.sheet.cell(b,3).value
    #         # old_data_cases.append(old_data)
    #         sheet3.cell(b,2).value = old_data
    # wb.save(url)





