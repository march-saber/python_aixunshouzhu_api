import openpyxl
class DoExcel:
    def __init__(self,excel_name,sheet_name):
        self.workbook = openpyxl.load_workbook(excel_name)
        self.sheet = self.workbook[sheet_name]

    def get_case(self):
        max_row = self.sheet.max_row  # 最大行数
        # max_col = self.sheet.max_column  #最大列数

        # case = []
        case_set = set()
        for r in range(2,max_row+1):
            case_time = self.sheet.cell(r,1).value
            # case_data = self.sheet.cell(r,2).value
            case_set.add(case_time)
        return case_set

    def get_time(self):
        max_row = self.sheet.max_row
        case = []
        a=0
        for r in range(2,max_row+1):
            case_time = self.sheet.cell(r,1).value
            case.append(case_time)
        for i in case:
            if i !=None:
                a+=1
        return a

    def get_newtime(self):
        max_row = self.sheet.max_row
        case = []
        a= 0
        for r in range(2, max_row+1):
            case_time = self.sheet.cell(r,2).value
            case.append(case_time)
        for i in case:
            if i !=None:
                a+=1
        return a