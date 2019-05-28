import openpyxl
import re

class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet = self.workbook[sheet_name]

    def get_case(self):
        max_row = self.sheet.max_row #最大行数
        max_col = self.sheet.max_column  #最大列数
        p = "(-)[1-9]*"


        cases = []
        line = 1
        for r in range(2,max_row+1):
            #for c in range(1,max_col+1):
            line_number = self.sheet.cell(r,1).value
            line += 1
            try:
                m = re.search(p,line_number)
                g = m.group(0)
            except AttributeError as e:
                print('第{0}行数据为空,日志{1}'.format(line,e))
            #去掉信号好强度：特点为负数
            # line_number = line_number.replace(g,' ')
            line_number = re.sub(g, ' ', line_number)

            cases.append(line_number)
            self.workbook.close()
        return cases

if __name__ == '__main__':
    from openpyxl import load_workbook
    do_excel1 = DoExcel('C:\\Users\\t\Desktop\数据记录\WiFi.xlsx','老版本第一次')
    do_excel2 = DoExcel('C:\\Users\\t\Desktop\数据记录\WiFi.xlsx','Pro第一次')

    cases1 = do_excel1.get_case()
    cases2 = do_excel2.get_case()
    # cases3 = []
    # cases4 = []
    wb = load_workbook('C:\\Users\\t\Desktop\数据记录\WiFi.xlsx')
    sheet3 = wb['交集']
    sheet4 = wb['差集']
    a = 0
    for item in cases1:
        if item in cases2:
            # cases3.append(item)  #交集
            a += 1
            sheet3.cell(a,1).value = item
            wb.save('C:\\Users\\t\Desktop\数据记录\WiFi.xlsx')
    #差集
    for item in cases1:
        if item not in cases2:
            # cases4.append(item)
            a += 1
            sheet4.cell(a, 1).value = item
            wb.save('C:\\Users\\t\Desktop\数据记录\WiFi.xlsx')
    for item in cases2:
        if item not in cases1:
            # cases4.append(item)
            a += 1
            sheet4.cell(a, 1).value = item
            wb.save('C:\\Users\\t\Desktop\数据记录\WiFi.xlsx')




